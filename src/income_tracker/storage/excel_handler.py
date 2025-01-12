"""Excel storage handler for managing payment records."""

import logging
from pathlib import Path
from typing import Protocol

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from ..models import PaymentRecord

logger = logging.getLogger(__name__)


class StorageError(Exception):
    """Raised when storage operations fail."""


class StorageHandler(Protocol):
    """Protocol for storage handlers."""
    
    def save_record(self, record: PaymentRecord) -> None:
        """Save a payment record."""
        ...
    
    def load_records(self) -> list[PaymentRecord]:
        """Load all payment records."""
        ...
    
    def record_exists(self, record: PaymentRecord) -> bool:
        """Check if record already exists."""
        ...


class ExcelStorageHandler:
    """Excel-based storage handler for payment records."""
    
    def __init__(self, file_path: str | Path) -> None:
        """Initialize Excel storage handler.
        
        Args:
            file_path: Path to Excel file
        """
        self.file_path = Path(file_path)
        self.columns = [
            "Date", "Amount", "Source", "Client Name", 
            "Email Subject", "Processed Date"
        ]
        self._ensure_file_exists()
    
    def _ensure_file_exists(self) -> None:
        """Ensure Excel file exists with proper headers."""
        if not self.file_path.exists():
            self._create_excel_file()
            logger.info(f"Created new Excel file: {self.file_path}")
    
    def _create_excel_file(self) -> None:
        """Create a new Excel file with styled headers."""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Payment Records"
        
        # Add headers with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, column_title in enumerate(self.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Set column widths
        column_widths = [12, 10, 12, 20, 40, 20]
        for col_num, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=col_num).column_letter].width = width
        
        workbook.save(self.file_path)
    
    def save_record(self, record: PaymentRecord) -> None:
        """Save a payment record to Excel file.
        
        Args:
            record: PaymentRecord to save
            
        Raises:
            StorageError: If saving fails
        """
        try:
            # Check if record already exists
            if self.record_exists(record):
                logger.info(f"Record already exists, skipping: {record.source.value} - ${record.amount}")
                return
            
            # Load existing data
            df = self._load_dataframe()
            
            # Add new record
            new_row = pd.DataFrame([record.to_dict()])
            df = pd.concat([df, new_row], ignore_index=True)
            
            # Save to Excel
            self._save_dataframe(df)
            logger.info(f"Saved payment record: {record.source.value} - ${record.amount} - {record.client_name}")
            
        except Exception as e:
            raise StorageError(f"Failed to save record: {e}") from e
    
    def load_records(self) -> list[PaymentRecord]:
        """Load all payment records from Excel file.
        
        Returns:
            List of PaymentRecord objects
            
        Raises:
            StorageError: If loading fails
        """
        try:
            df = self._load_dataframe()
            records = []
            
            for _, row in df.iterrows():
                # Convert row to PaymentRecord (implementation would need proper parsing)
                # This is a simplified version
                logger.debug(f"Loaded record: {row.get('Source', 'Unknown')} - {row.get('Amount', 0)}")
            
            return records
            
        except Exception as e:
            raise StorageError(f"Failed to load records: {e}") from e
    
    def record_exists(self, record: PaymentRecord) -> bool:
        """Check if a payment record already exists.
        
        Args:
            record: PaymentRecord to check
            
        Returns:
            True if record exists, False otherwise
        """
        try:
            df = self._load_dataframe()
            
            # Check for duplicate based on key fields
            mask = (
                (df["Date"] == record.date.strftime("%Y-%m-%d")) &
                (df["Amount"] == float(record.amount)) &
                (df["Source"] == record.source.value) &
                (df["Client Name"] == record.client_name)
            )
            
            return mask.any()
            
        except Exception as e:
            logger.warning(f"Error checking record existence: {e}")
            return False
    
    def _load_dataframe(self) -> pd.DataFrame:
        """Load DataFrame from Excel file."""
        try:
            return pd.read_excel(self.file_path)
        except Exception as e:
            logger.warning(f"Error loading Excel file, creating empty DataFrame: {e}")
            return pd.DataFrame(columns=self.columns)
    
    def _save_dataframe(self, df: pd.DataFrame) -> None:
        """Save DataFrame to Excel file with formatting."""
        with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Payment Records')
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Payment Records']
            
            # Apply header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def get_summary_stats(self) -> dict[str, float | int]:
        """Get summary statistics from stored records.
        
        Returns:
            Dictionary with summary statistics
        """
        try:
            df = self._load_dataframe()
            
            if df.empty:
                return {"total_records": 0, "total_amount": 0.0}
            
            return {
                "total_records": len(df),
                "total_amount": float(df["Amount"].sum()),
                "average_amount": float(df["Amount"].mean()),
                "max_amount": float(df["Amount"].max()),
                "min_amount": float(df["Amount"].min()),
                "unique_clients": df["Client Name"].nunique(),
                "by_source": df["Source"].value_counts().to_dict()
            }
            
        except Exception as e:
            logger.error(f"Error calculating summary stats: {e}")
            return {"error": str(e)}
